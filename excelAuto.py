#import pandas as pd
import openpyxl
from openpyxl import load_workbook, worksheet, Workbook
from openpyxl.styles import Font, Color
from openpyxl.worksheet.formula import ArrayFormula

wb = openpyxl.load_workbook('ProfileVV.xlsx')
target = wb['Variable vs Variable']
wb.copy_worksheet(target)
print(wb.sheetnames)
#apply font and color
ft=Font(color="FF0000")
ws=wb['Variable vs Variable Copy']
ws['K4'] = "%Pen"
ws['L4'] = "Index"
col1 = ws.column_dimensions['K']
col1.font=ft
col2 = ws.column_dimensions['L']
col2.font=ft
# Apply formula
ws["K3"] = ArrayFormula("K5:K72", "=E5:E72/G5:G72*100")    
ws["L3"] = ArrayFormula("L5:L72", "=F5:F72/H5:H72*100")


wb.save('ProfileVV1.xlsx')
#for row in ws.iter_cols(min_col=10, max_col=10, min_row=4,max_row=30):
 #   for cell in row:
  #      print(cell.value)


for x in range(4,30):
    count = ws.cell(row=x, column=5).value
    base_count = ws.cell(row=x, column=7).value
    index = ws.cell(row=x, column=10).value 
    print(count,base_count,index)

